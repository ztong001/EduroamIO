from openpyxl import *
from IHL import *

try:
	wb= load_workbook('Eduroam test stats.xlsx')
except FileNotFoundError:
	wb= Workbook()
	## First worksheet of the spreadsheet
	ws= wb.active
	## Second worksheet of the spreadsheet
	ws1= wb.create_sheet()
	ws.title = "Daily"
	ws['A2']= 'Date'
	ws['B1']= 'NUS'
	ws['B2']= 'Local Users'
	ws['C2']= 'Visitors'
	ws['D2']= 'Unsuccessful Attempts'
	ws['E2']= 'Date'
	ws['F1']= 'NTU'
	ws['F2']= 'Local Users'
	ws['G2']= 'Visitors'
	ws['H2']= 'Unsuccessful Attempts'
	ws['I2']= 'Date'
	ws['J1']= 'SMU'
	ws['J2']= 'Local Users'
	ws['K2']= 'Visitors'
	ws['L2']= 'Unsuccessful Attempts'
	ws['M2']= 'Date'
	ws['N1']= 'ASTAR'
	ws['N2']= 'Local Users'
	ws['O2']= 'Visitors'
	ws['P2']= 'Unsuccessful Attempts'
	ws['Q2']= 'Date'
	ws['R1']= 'NIE'
	ws['R2']= 'Local Users'
	ws['S2']= 'Visitors'
	ws['T2']= 'Unsuccessful Attempts'

	ws1.title = "Monthly"
	ws1['A2']= 'Month'
	ws1['B1']= 'NUS'
	ws1['B2']= 'UniqueUsers'
	ws1['C2']= 'Unsuccessful Attempts'
	ws1['D2']= 'Month'
	ws1['E1']= 'NTU'
	ws1['E2']= 'UniqueUsers'
	ws1['F2']= 'Unsuccessful Attempts'
	ws1['G2']= 'Month'
	ws1['H1']= 'SMU'
	ws1['H2']= 'UniqueUsers'
	ws1['I2']= 'Unsuccessful Attempts'
	ws1['J2']= 'Month'
	ws1['K1']= 'ASTAR'
	ws1['K2']= 'UniqueUsers'
	ws1['L2']= 'Unsuccessful Attempts'
	ws1['M2']= 'Month'
	ws1['N1']= 'NIE'
	ws1['N2']= 'UniqueUsers'
	ws1['O2']= 'Unsuccessful Attempts'

	print("Created new Worksheet!")
	wb.save('Eduroam test stats.xlsx')
	print("Created new Excel file!")
else:
	print("Opening Eduroam test stats!")

#### Entering Test Values for Excel ####
nus= IHL("NUS")
ntu= IHL("NTU")
smu= IHL("SMU")
astar= IHL("ASTAR")
nie= IHL("NIE")
IHL_Array=(nus,ntu,smu,astar,nie)
for ihl in IHL_Array:
	ihl.localUsers_abroad= 1
	ihl.localUsers_astar= 2
	ihl.localUsers_nie= 3
	ihl.localUsers_ntu= 4
	ihl.localUsers_nus= 5
	ihl.localUsers_smu= 6
	ihl.localUsers= ihl.localUsers_abroad+ ihl.localUsers_astar+ ihl.localUsers_nie+ ihl.localUsers_ntu+ ihl.localUsers_nus+ ihl.localUsers_smu
	ihl.reject_localUsers=5
	ihl.reject_visitors=5
	ihl.visitors=4
	ihl.uniqueUsersMonth= 13
	ihl.rejectUniqueUsersMonth= 3

#### Entering Data in Daily Sheet ####
## A3= First Date ##
dCell= dCell2= 3
ws= wb.get_sheet_by_name("Daily")
ws1= wb.get_sheet_by_name("Monthly")
ws['A'+str(dCell)]='280515'
ws['E'+str(dCell)]='280515'
ws['I'+str(dCell)]='280515'
ws['M'+str(dCell)]='280515'
ws['Q'+str(dCell)]='280515'
#NUS
ws['B'+str(dCell)]= nus.localUsers
ws['C'+str(dCell)]= nus.visitors
ws['D'+str(dCell)]= nus.getRejectCount()
#NTU
ws['F'+str(dCell)]= ntu.localUsers
ws['G'+str(dCell)]= ntu.visitors
ws['H'+str(dCell)]= ntu.getRejectCount()
#SMU
ws['J'+str(dCell)]= smu.localUsers
ws['K'+str(dCell)]= smu.visitors
ws['L'+str(dCell)]= smu.getRejectCount()

#### Entering Data in Monthly Sheet ####
ws1['A'+str(dCell2)]='MAY'
ws1['D'+str(dCell2)]='MAY'
ws1['G'+str(dCell2)]='MAY'
ws1['J'+str(dCell2)]='MAY'
ws1['M'+str(dCell2)]='MAY'
#NUS
ws1['B'+str(dCell2)]= nus.uniqueUsersMonth
ws1['C'+str(dCell2)]= nus.rejectUniqueUsersMonth
#NTU
ws1['E'+str(dCell2)]= ntu.uniqueUsersMonth
ws1['F'+str(dCell2)]= ntu.rejectUniqueUsersMonth
#SMU
ws1['H'+str(dCell2)]= smu.uniqueUsersMonth
ws1['I'+str(dCell2)]= smu.rejectUniqueUsersMonth

print('Completed update!')
wb.save('Eduroam test stats.xlsx')
print('Workbook saved!')